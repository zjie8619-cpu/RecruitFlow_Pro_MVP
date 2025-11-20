from __future__ import annotations

from io import BytesIO
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill


HEADER_COLUMNS = [
    "能力维度",
    "说明",
    "权重（%）",
    "面试问题",
    "评分要点",
    "20分行为表现",
    "60分行为表现",
    "100分行为表现",
]


def _format_points(points: Any) -> str:
    if isinstance(points, list):
        return "\n".join(str(p) for p in points if str(p).strip())
    if points is None:
        return ""
    return str(points)


def _format_question(question: Any) -> str:
    if isinstance(question, list):
        return "\n".join(str(q) for q in question if str(q).strip())
    if question is None:
        return ""
    return str(question)


def generate_competency_excel(
    competencies: List[Dict[str, Any]],
    interview_questions: List[Dict[str, Any]],
    file_path: Optional[str] = None,
) -> bytes:
    """
    生成岗位能力维度评分的 Excel，并返回二进制内容。

    Args:
        competencies: 维度列表（每项包含 name/desc/weight/anchors 等字段）
        interview_questions: 面试题列表（每项包含 dimension/question/points 等）
        file_path: 若提供，将额外将 Excel 保存到指定路径

    Returns:
        bytes: Excel 文件二进制内容，可用于下载
    """
    question_map = {q.get("dimension"): q for q in interview_questions or []}

    wb = Workbook()
    ws = wb.active
    ws.title = "能力维度评分"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 写入表头
    for col_idx, title in enumerate(HEADER_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = thin_border
        cell.fill = header_fill

    # 写入数据行
    for row_idx, dim in enumerate(competencies or [], start=2):
        anchors = dim.get("anchors") or {}
        question_entry = question_map.get(dim.get("name")) or {}
        question_text = _format_question(question_entry.get("question"))
        points_text = _format_points(question_entry.get("points"))

        row_values = [
            dim.get("name", ""),
            dim.get("desc", ""),
            round(float(dim.get("weight", 0.0)) * 100, 1),
            question_text,
            points_text,
            anchors.get("20", ""),
            anchors.get("60", ""),
            anchors.get("100", ""),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = alignment_left if col_idx != 3 else alignment_center

    # 自动列宽设置
    width_map = {
        "A": 18,
        "B": 36,
        "C": 12,
        "D": 36,
        "E": 32,
        "F": 30,
        "G": 30,
        "H": 30,
    }
    for col_letter, width in width_map.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 28

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    data = buffer.getvalue()

    if file_path:
        with open(file_path, "wb") as f:
            f.write(data)

    return data

