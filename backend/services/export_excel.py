import os
from io import BytesIO
from typing import Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEFAULT_TEMPLATE_PATH = r"C:\RecruitFlow_Pro_MVP\docs\课程顾问_能力维度评分表 (改).xlsx"
TEMPLATE_PATH = globals().get("TEMPLATE_PATH", DEFAULT_TEMPLATE_PATH)

COLUMN_ALIASES = {
    "dimension": "能力维度",
    "name": "能力维度",
    "ability": "能力维度",
    "能力维度": "能力维度",
    "说明": "说明",
    "description": "说明",
    "desc": "说明",
    "面试题目": "面试题目",
    "question": "面试题目",
    "interview_question": "面试题目",
    "评分要点": "评分要点",
    "score_points": "评分要点",
    "points": "评分要点",
    "20分行为表现": "20分行为表现",
    "score_20": "20分行为表现",
    "60分行为表现": "60分行为表现",
    "score_60": "60分行为表现",
    "100分行为表现": "100分行为表现",
    "score_100": "100分行为表现",
    "权重": "权重",
    "weight": "权重",
}

EXPECTED_COLUMNS = [
    "能力维度",
    "说明",
    "面试题目",
    "评分要点",
    "20分行为表现",
    "60分行为表现",
    "100分行为表现",
    "权重",
]


def _normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    renamed = {}
    for col in df.columns:
        key = str(col).strip()
        renamed[col] = COLUMN_ALIASES.get(key, key)
    normalized = df.rename(columns=renamed)
    for col in EXPECTED_COLUMNS:
        if col not in normalized.columns:
            normalized[col] = ""
    return normalized[EXPECTED_COLUMNS]


def _make_alt_path(original: str) -> str:
    base, ext = os.path.splitext(original)
    from datetime import datetime

    suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_{suffix}{ext}"


def _build_score_desc(row: pd.Series) -> str:
    """构建对应分值描述文本，格式：100分：...\n60分：...\n20分：..."""
    parts = []
    # 按100、60、20的顺序添加，确保格式一致
    score_100 = row.get("100分行为表现") or row.get("score_100") or ""
    score_60 = row.get("60分行为表现") or row.get("score_60") or ""
    score_20 = row.get("20分行为表现") or row.get("score_20") or ""
    
    if score_100:
        parts.append(f"100分：{str(score_100).strip()}")
    if score_60:
        parts.append(f"60分：{str(score_60).strip()}")
    if score_20:
        parts.append(f"20分：{str(score_20).strip()}")
    
    return "\n".join(parts) if parts else ""


def export_competency_excel(
    data_df: pd.DataFrame,
    output_path: Optional[str] = None,
    template_path: Optional[str] = None,
    job_title: Optional[str] = None,
) -> Tuple[bytes, Optional[str]]:
    """
    根据模板导出岗位能力维度评分表。
    返回 (excel_bytes, saved_path)。如保存失败则 saved_path 为 None。
    """

    if data_df is None or data_df.empty:
        raise ValueError("能力维度数据为空, 无法导出 Excel。")

    template_path = template_path or TEMPLATE_PATH
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"模板文件不存在: {template_path}. 请确认《能力维度评分表 (改).xlsx》已放在 docs 目录。"
        )

    data_df = _normalize_df(data_df)
    wb = load_workbook(template_path)
    ws = wb.active
    
    # ======== 从模板克隆样式，而不是手写样式 ========
    # 这样可以 100% 跟随模板的字号、对齐、边框、锁定状态
    #
    # 注意：在删除/插入行前，先把需要参考的“表头行”和“一行示例数据行”的样式保存下来
    
    # 边框样式（thin边框）
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 先查找表头行和合计总分行
    header_row = None
    total_row = None
    for row in ws.iter_rows(min_row=1, max_col=1):
        value = row[0].value
        if value == "序号" and header_row is None:
            header_row = row[0].row
        if isinstance(value, str) and "合计总分" in value:
            total_row = row[0].row
            break
    if header_row is None or total_row is None:
        raise ValueError("模板结构识别失败, 请确认模板包含\"序号\"和\"合计总分：\"行。")
    
    # ======== 保存模板示例样式（表头 + 第一条数据行） ========
    # 1）列宽（在删除/插入行之前保存）
    template_column_widths = {}
    for col in range(1, 10):
        col_letter = get_column_letter(col)
        if col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width
            if width:
                template_column_widths[col_letter] = width

    # 2）保存表头单元格样式
    header_styles = {}
    for col in range(1, 10):
        cell = ws.cell(row=header_row, column=col)
        header_styles[col] = cell._style

    # 3）保存“模板中的第一条数据行”的样式，用于克隆到所有数据行
    #    默认认为 header_row + 1 就是模板示例行
    template_data_row_index = header_row + 1
    data_row_styles = {}
    if template_data_row_index <= ws.max_row:
        for col in range(1, 10):
            cell = ws.cell(row=template_data_row_index, column=col)
            data_row_styles[col] = cell._style
    else:
        # 万一模板里没有示例行，则退回到仅使用边框，其他使用 openpyxl 默认
        for col in range(1, 10):
            dummy_cell = ws.cell(row=header_row, column=col)
            data_row_styles[col] = dummy_cell._style
    
    # 保存模板的行高（用于数据行）
    template_data_row_height = 57.0  # 默认数据行高
    # 尝试从模板中获取数据行高（同样参考 header_row+1 这一行）
    if template_data_row_index in ws.row_dimensions:
        height = ws.row_dimensions[template_data_row_index].height
        if height:
            template_data_row_height = height
    
    # 处理标题行
    if job_title:
        title_cell = ws["A1"]
        title_cell.value = f"{job_title}岗位能力维度与面试题目"
        # 确保标题行合并
        existing_merged = [str(m) for m in ws.merged_cells.ranges if m.min_row == 1]
        if 'A1:I1' not in existing_merged:
            # 先取消可能存在的旧合并
            for merged in list(ws.merged_cells.ranges):
                if merged.min_row == 1:
                        ws.unmerge_cells(str(merged))
                ws.merge_cells('A1:I1')
        # 标题本身的样式直接沿用模板中 A1 的原始样式，只修改文本即可
        # 若用户在模板中调整字号/加粗/对齐等，这里会自动跟随
        # 同时保留原有行高

    data_start_row = header_row + 1
    existing_rows = max(total_row - data_start_row, 0)
    if existing_rows:
        ws.delete_rows(data_start_row, existing_rows)

    ws.insert_rows(data_start_row, len(data_df))

    # 应用数据行格式
    for idx, (_, row) in enumerate(data_df.iterrows(), start=1):
        excel_row = data_start_row + idx - 1
        
        # 设置行高（按模板示例行高度）
        ws.row_dimensions[excel_row].height = template_data_row_height
        
        # 先按列依次写入内容 & 克隆样式（完全跟随模板示例行）
        # 1 列：序号
        cell = ws.cell(excel_row, 1, idx)
        if 1 in data_row_styles:
            cell._style = data_row_styles[1]
        cell.border = thin_border
        
        # 2 列：能力维度
        cell = ws.cell(excel_row, 2, row["能力维度"])
        if 2 in data_row_styles:
            cell._style = data_row_styles[2]
        cell.border = thin_border
        
        # 3 列：说明
        cell = ws.cell(excel_row, 3, row["说明"])
        if 3 in data_row_styles:
            cell._style = data_row_styles[3]
        cell.border = thin_border
        
        # 4 列：面试题目
        cell = ws.cell(excel_row, 4, row["面试题目"])
        if 4 in data_row_styles:
            cell._style = data_row_styles[4]
        cell.border = thin_border
        
        # 5 列：评分要点
        cell = ws.cell(excel_row, 5, row["评分要点"])
        if 5 in data_row_styles:
            cell._style = data_row_styles[5]
        cell.border = thin_border
        
        # 6 列：对应分值描述（由 20/60/100 行为表现拼接而来）
        score_desc = _build_score_desc(row)
        cell = ws.cell(excel_row, 6, score_desc)
        if 6 in data_row_styles:
            cell._style = data_row_styles[6]
        cell.border = thin_border
        
        # 7 列：权重
        weight_value = row["权重"]
        if isinstance(weight_value, (int, float)):
            weight_value = float(weight_value)
        cell = ws.cell(excel_row, 7, weight_value)
        if 7 in data_row_styles:
            cell._style = data_row_styles[7]
        cell.border = thin_border
        
        # 8 列：评分（空，由面试官填写）
        cell = ws.cell(excel_row, 8, None)
        if 8 in data_row_styles:
            cell._style = data_row_styles[8]
        cell.border = thin_border
        
        # 9 列：得分（空，由面试官填写/公式计算）
        cell = ws.cell(excel_row, 9, None)
        if 9 in data_row_styles:
            cell._style = data_row_styles[9]
        cell.border = thin_border
    
    # 确保表头行格式正确：按模板表头克隆样式
    for col in range(1, 10):
        cell = ws.cell(header_row, col)
        if col in header_styles:
            cell._style = header_styles[col]
        cell.border = thin_border
    
    # 确保合计总分行格式正确
    if total_row:
        total_cell = ws.cell(total_row, 1)
        if total_cell.value and isinstance(total_cell.value, str) and "合计总分" in total_cell.value:
            # 确保合计总分行合并（A到H列）
            total_range = f"A{total_row}:H{total_row}"
            existing_merged = [str(m) for m in ws.merged_cells.ranges if m.min_row == total_row]
            if total_range not in existing_merged:
                # 先取消可能存在的旧合并
                for merged in list(ws.merged_cells.ranges):
                    if merged.min_row == total_row:
                        ws.unmerge_cells(str(merged))
                ws.merge_cells(total_range)
            # 合计总分行的样式尽量保持模板原样，仅补边框
            total_cell.border = thin_border
            # 确保合计总分行 I 列也有边框（其他样式仍沿用模板）
            cell_i = ws.cell(total_row, 9)
            cell_i.border = thin_border
    
    # 恢复并锁定模板的列宽（使用默认值确保列宽正确）
    default_widths = {
        'A': 2.90740740740741,
        'B': 5.74074074074074,
        'C': 8.57407407407407,
        'D': 8.5,
        'E': 8.11111111111111,
        'F': 21.9074074074074,
        'G': 6.96296296296296,
        'H': 13.0,
        'I': 13.8703703703704
    }
    
    # 优先使用模板中的列宽，如果没有则使用默认值
    for col_letter, default_width in default_widths.items():
        if col_letter in template_column_widths and template_column_widths[col_letter]:
            ws.column_dimensions[col_letter].width = template_column_widths[col_letter]
        else:
            ws.column_dimensions[col_letter].width = default_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    saved_path = None
    if output_path:
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with open(output_path, "wb") as f:
                f.write(buffer.getvalue())
            saved_path = output_path
        except PermissionError:
            alt_path = _make_alt_path(output_path)
            with open(alt_path, "wb") as f:
                f.write(buffer.getvalue())
            saved_path = alt_path

    return buffer.getvalue(), saved_path
