from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def describe_cell(cell):
    return {
        "font_size": cell.font.sz,
        "bold": cell.font.bold,
        "name": cell.font.name,
        "h_align": cell.alignment.horizontal,
        "v_align": cell.alignment.vertical,
        "wrap": cell.alignment.wrap_text,
    }


def main():
    base = Path(r"C:\RecruitFlow_Pro_MVP")
    tpl_path = base / "docs" / "课程顾问_能力维度评分表 (改).xlsx"
    out_path = base / "docs" / "课程顾问_能力维度评分表(改)_输出.xlsx"

    print("TEMPLATE exists:", tpl_path.exists(), "->", tpl_path)
    print("OUTPUT   exists:", out_path.exists(), "->", out_path)

    wb_tpl = load_workbook(tpl_path)
    ws_tpl = wb_tpl.active
    wb_out = load_workbook(out_path)
    ws_out = wb_out.active

    header_row = None
    total_row = None
    for row in ws_tpl.iter_rows(min_row=1, max_col=1):
        v = row[0].value
        if v == "序号" and header_row is None:
            header_row = row[0].row
        if isinstance(v, str) and "合计总分" in v:
            total_row = row[0].row
            break

    print("header_row =", header_row, " total_row =", total_row)

    if not header_row:
        print("!! cannot find header_row in template")
        return

    first_data_row_tpl = header_row + 1
    first_data_row_out = header_row + 1

    print("\n=== HEADER COMPARISON (TEMPLATE vs OUTPUT) ===")
    for col in range(1, 10):
        lt = get_column_letter(col)
        ct = ws_tpl.cell(header_row, col)
        co = ws_out.cell(header_row, col)
        dt = describe_cell(ct)
        do = describe_cell(co)
        print(
            f"Col {lt}: tpl={dt} | out={do}"
        )

    print("\n=== FIRST DATA ROW COMPARISON (TEMPLATE vs OUTPUT) ===")
    print("tpl row =", first_data_row_tpl, " out row =", first_data_row_out)
    for col in range(1, 10):
        lt = get_column_letter(col)
        ct = ws_tpl.cell(first_data_row_tpl, col)
        co = ws_out.cell(first_data_row_out, col)
        dt = describe_cell(ct)
        do = describe_cell(co)
        print(
            f"Col {lt}: tpl={dt} | out={do}"
        )


if __name__ == "__main__":
    main()


